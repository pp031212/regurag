"""XLSX 文本抽取服务。

每个工作表按行展开为管道分隔文本，方便后续分块和关键词检索。
"""

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..pdf.utils import table_rows_to_text


@dataclass(slots=True)
class XLSXTextExtractionArtifacts:
    """XLSX 预处理产物路径。"""

    extracted_txt_path: Path


class XLSXTextExtractionService:
    """提取 Excel 工作簿中的可检索文本。"""

    def preprocess(self, xlsx_path: Path, output_dir: Path) -> XLSXTextExtractionArtifacts:
        """生成 extracted.txt。"""
        output_dir.mkdir(parents=True, exist_ok=True)
        extracted_txt_path = output_dir / "extracted.txt"

        # data_only=True 读取公式缓存值，避免把公式表达式直接入库。
        workbook = load_workbook(xlsx_path, data_only=True)
        lines: list[str] = []

        for sheet in workbook.worksheets:
            lines.append(f"[Sheet] {sheet.title}")
            merged_values = self._build_merged_cell_value_map(sheet)
            sheet_rows: list[list[str]] = []
            for row in sheet.iter_rows():
                cells: list[str] = []
                for cell in row:
                    value = merged_values.get(cell.coordinate, cell.value)
                    if value is not None and str(value).strip():
                        cells.append(str(value).strip())
                if cells:
                    sheet_rows.append(cells)
            sheet_text = table_rows_to_text(sheet_rows)
            if sheet_text:
                lines.append(sheet_text)

        extracted_text = "\n\n".join(lines).strip()
        extracted_txt_path.write_text(extracted_text, encoding="utf-8")

        return XLSXTextExtractionArtifacts(extracted_txt_path=extracted_txt_path)

    @staticmethod
    def _build_merged_cell_value_map(sheet: Any) -> dict[str, object]:
        """把合并单元格的左上角值填充到整个合并区域，保留父级表头语义。"""
        merged_values: dict[str, object] = {}
        for merged_range in sheet.merged_cells.ranges:
            top_left = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
            if top_left is None or not str(top_left).strip():
                continue
            for row in sheet.iter_rows(
                min_row=merged_range.min_row,
                max_row=merged_range.max_row,
                min_col=merged_range.min_col,
                max_col=merged_range.max_col,
            ):
                for cell in row:
                    merged_values[cell.coordinate] = top_left
        return merged_values
