from pathlib import Path

from openpyxl import Workbook

from app.document_processing.xlsx import XLSXTextExtractionService


def test_xlsx_text_extraction_service_exports_sheet_names_and_rows(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "sample.xlsx"
    output_dir = tmp_path / "artifacts"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "处分表"
    sheet.append(["行为", "扣分"])
    sheet.append(["旷课", 20])
    workbook.create_sheet("说明")
    workbook["说明"].append(["备注", "按周累计"])
    workbook.save(xlsx_path)

    service = XLSXTextExtractionService()
    artifacts = service.preprocess(xlsx_path, output_dir)

    extracted_text = artifacts.extracted_txt_path.read_text(encoding="utf-8")
    assert "[Sheet] 处分表" in extracted_text
    assert "行为 | 扣分" in extracted_text
    assert "旷课 | 20" in extracted_text
    assert "行为: 旷课" in extracted_text
    assert "扣分: 20" in extracted_text
    assert "[Sheet] 说明" in extracted_text
    assert "备注 | 按周累计" in extracted_text


def test_xlsx_text_extraction_service_expands_merged_cells(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "merged.xlsx"
    output_dir = tmp_path / "artifacts"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "扣分表"
    sheet["A1"] = "课堂纪律"
    sheet.merge_cells("A1:A3")
    sheet["B1"] = "迟到"
    sheet["C1"] = "扣3分/次"
    sheet["B2"] = "旷课"
    sheet["C2"] = "扣20分/次"
    sheet["B3"] = "早退"
    sheet["C3"] = "扣3分/次"
    workbook.save(xlsx_path)

    service = XLSXTextExtractionService()
    artifacts = service.preprocess(xlsx_path, output_dir)

    extracted_text = artifacts.extracted_txt_path.read_text(encoding="utf-8")
    assert "课堂纪律 | 迟到 | 扣3分/次" in extracted_text
    assert "课堂纪律 | 旷课 | 扣20分/次" in extracted_text
    assert "课堂纪律 | 早退 | 扣3分/次" in extracted_text
